import json

from openpyxl import load_workbook

from src.develop.export_l2_review_workbook import export_workbook


def _contract():
    return {
        "contract_version": "l2_sentence_regions_v3",
        "artifact_status": "DRAFT_NEEDS_HUMAN_ADJUDICATION",
        "review_rows": 1,
        "review_reason_counts": {"MULTI_INDICATOR_BOUNDARY": 1},
        "id_uniqueness_scope": "article",
        "scope_id_format": "{article_idx}-SC{nn}",
        "region_id_format": "{article_idx}-R{nn}",
        "source_region_subtypes": ["공식집계", "민간조사"],
        "label_provenance_values": ["UNREVIEWED", "HUMAN_CONFIRMED"],
        "span_resolution_rule": "text only",
        "boundary_reference_rule": "span ids only",
        "reference_integrity_rule": "same article",
        "disagreement_rule": "note required",
    }


def _human():
    return [{
        "sentence_review_id": "1-S001",
        "article_idx": "1",
        "text": "빵 물가 상승률은 38.5%다.",
        "value_candidate_span_ids": "38.5%=s1:value_unit:2-7",
        "review_reason": "MULTI_INDICATOR_BOUNDARY",
        "indicator_scopes_json": "",
        "source_regions_json": "",
        "period_contexts_json": "",
        "clause_value_boundaries_json": "",
        "dominant_region_decision": "",
        "label_provenance": "UNREVIEWED",
        "review_status": "미검토",
        "reviewer_note": "",
    }]


def _context():
    return [{
        "sentence_review_id": "1-S000",
        "article_idx": "1",
        "sentence_id": 0,
        "text": "리드 문장.",
        "row_kind": "자동확정",
        "region_id": "1-R01",
        "scope_id": "1-SC01",
        "indicator_label": "빵 물가 상승률",
        "source_subtype": "공식집계",
        "derivation_status": "SINGLE_INDICATOR_EXPLICIT",
        "disagree_flag": "",
        "reviewer_note": "",
    }]


def test_workbook_has_three_sheets_and_no_offset_columns(tmp_path):
    output = tmp_path / "review.xlsx"

    result = export_workbook(_human(), _context(), _contract(), output)

    assert output.exists()
    assert result["offset_columns_present"] is False
    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "검토입력",
        "기사별 문맥 117문장",
        "계약·가이드",
    ]
    header = [cell.value for cell in workbook["검토입력"][1]]
    assert not [name for name in header if "char" in str(name)]
    assert "value_candidate_span_ids" in header


def test_workbook_preserves_empty_human_inputs(tmp_path):
    output = tmp_path / "review.xlsx"

    export_workbook(_human(), _context(), _contract(), output)

    sheet = load_workbook(output)["검토입력"]
    header = [cell.value for cell in sheet[1]]
    row = dict(zip(header, [cell.value for cell in sheet[2]]))
    for field in (
        "indicator_scopes_json",
        "source_regions_json",
        "period_contexts_json",
        "clause_value_boundaries_json",
        "dominant_region_decision",
    ):
        assert row[field] in (None, "")
    assert row["label_provenance"] == "UNREVIEWED"


def test_workbook_guide_states_offsets_are_derived(tmp_path):
    output = tmp_path / "review.xlsx"

    export_workbook(_human(), _context(), _contract(), output)

    guide = load_workbook(output)["계약·가이드"]
    rendered = json.dumps(
        [[cell.value for cell in row] for row in guide.iter_rows()],
        ensure_ascii=False,
    )
    assert "char offset" in rendered
    assert "occurrence_index" in rendered
