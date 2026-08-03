"""Consolidate L-STRATUM judgments into one freeze-ready sheet.

Judgments arrived in two passes (the first 45-article sheet and the re-check
workbook).  Two detector defects were fixed between them — a Korean
organisation name embedded in a longer word and a foreign ministry sharing a
Korean name — so an article judged earlier may no longer be a candidate at all.

Articles are kept only when the current detector still finds the organisation
and the body actually attributes a figure to a statistic it publishes.  Naming
an organisation as an actor is not that evidence, and an article without a
KOSIS statistic would end as ``UNVERIFIABLE`` regardless of how well the
structuring layer performed, which would depress Gate B recall for a reason
that has nothing to do with the layer under test.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

KOSIS = "KOSIS등재"
OUTPUT_COLUMNS = (
    "article_idx",
    "기사제목",
    "작성일",
    "current_gold_source_scope",
    "value_candidate_count",
    "density_bin",
    "judged_source_scope",
    "judge_note",
)


def _read_jsonl(path: Path) -> dict[str, dict[str, Any]]:
    return {
        json.loads(line)["article_idx"]: json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    }


def _read_sheet(path: Path, sheet: str, header_row: int) -> dict[str, str]:
    worksheet = load_workbook(path)[sheet]
    header = [cell.value for cell in worksheet[header_row]]
    i_article = header.index("article_idx")
    i_judged = header.index("judged_source_scope")
    judgments: dict[str, str] = {}
    for index in range(header_row + 1, worksheet.max_row + 1):
        article_idx = worksheet.cell(index, i_article + 1).value
        if not article_idx:
            continue
        judgments[str(article_idx)] = str(
            worksheet.cell(index, i_judged + 1).value or ""
        ).strip()
    return judgments


def consolidate(
    judgments: dict[str, str],
    pool: dict[str, dict[str, Any]],
    *,
    require_source_shaped: bool = True,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    accepted: list[dict[str, Any]] = []
    dropped_not_candidate: list[str] = []
    dropped_no_source_evidence: list[str] = []
    non_kosis: list[str] = []
    for article_idx, verdict in judgments.items():
        if verdict != KOSIS:
            non_kosis.append(article_idx)
            continue
        row = pool.get(article_idx)
        if row is None:
            dropped_not_candidate.append(article_idx)
            continue
        summary = row.get("org_role_summary") or {}
        if require_source_shaped and not summary.get("source_shaped"):
            dropped_no_source_evidence.append(article_idx)
            continue
        record = dict(row)
        record["judged_source_scope"] = KOSIS
        accepted.append(record)
    accepted.sort(key=lambda row: int(row["article_idx"]))
    report = {
        "judged_rows": len(judgments),
        "judged_kosis": sum(1 for v in judgments.values() if v == KOSIS),
        "accepted": len(accepted),
        "dropped_detector_false_positive": sorted(
            dropped_not_candidate, key=int
        ),
        "dropped_no_source_shaped_evidence": sorted(
            dropped_no_source_evidence, key=int
        ),
        "judged_non_kosis": sorted(non_kosis, key=int),
        "require_source_shaped": require_source_shaped,
        "density_counts": pd.Series(
            [row["density_bin"] for row in accepted]
        ).value_counts().to_dict() if accepted else {},
    }
    return accepted, report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--first-pass", type=Path, required=True)
    parser.add_argument("--recheck", type=Path, required=True)
    parser.add_argument("--pool", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--output-internal", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--allow-actor-only", action="store_true")
    args = parser.parse_args()

    judgments = _read_sheet(args.first_pass, "판정입력", 4)
    for sheet in ("재확인", "신규후보"):
        judgments.update(_read_sheet(args.recheck, sheet, 1))

    accepted, report = consolidate(
        judgments,
        _read_jsonl(args.pool),
        require_source_shaped=not args.allow_actor_only,
    )

    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        [{name: row.get(name) for name in OUTPUT_COLUMNS} for row in accepted]
    ).to_csv(args.output_csv, index=False, encoding="utf-8-sig")
    args.output_internal.write_text(
        "".join(
            json.dumps(row, ensure_ascii=False) + "\n" for row in accepted
        ),
        encoding="utf-8",
    )
    args.report.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
