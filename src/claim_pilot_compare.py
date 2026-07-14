"""
Task 3 (claim_class 라벨링 파일럿) 2단계: 사람 라벨 vs LLM 판단 비교.

`claim_pilot_sample.py`로 만든 data/claim_class_labeling_pilot.xlsx의
'라벨링' 시트에서 claim_class/source_scope를 사람이 채운 뒤 실행한다.

주의: 문장 하나당 HCX API를 1회 호출한다 (비용 발생). 30~50건이면 미미하지만
반복 실행하지 말 것 — 이미 llm_verdict가 채워진 행은 건너뛴다(재실행해도 안전).

사용 예 (레포 루트에서, 라벨링 끝난 뒤):
    venv/Scripts/python.exe src/claim_pilot_compare.py
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from llm_claim_extractor import extract_claims_from_article, NCP_MODEL  # noqa: E402

WORKBOOK_PATH = Path("data/claim_class_labeling_pilot.xlsx")
REPORT_PATH = Path("docs/claim_class_labeling_pilot.md")


def human_verifiable(claim_class: str, source_scope: str) -> bool:
    return claim_class == "집계통계" and source_scope == "KOSIS계열"


def run_llm_judgement(claim_text: str) -> bool:
    """이 문장을 '기사'로 취급해 LLM에게 넣고, 하나라도 claim으로 추출되면 True."""
    claims, _usage = extract_claims_from_article(claim_text)
    return len(claims) > 0


def main() -> None:
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8")

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["라벨링"]
    headers = [c.value for c in ws[1]]

    if "llm_verifiable" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="llm_verifiable")
        headers.append("llm_verifiable")

    col = {name: i + 1 for i, name in enumerate(headers)}
    unlabeled, missing_llm = 0, 0

    for r in range(2, ws.max_row + 1):
        claim_class = ws.cell(row=r, column=col["claim_class"]).value
        source_scope = ws.cell(row=r, column=col["source_scope"]).value
        claim_text = ws.cell(row=r, column=col["claim_text"]).value
        llm_value = ws.cell(row=r, column=col["llm_verifiable"]).value

        if not claim_class or not source_scope:
            unlabeled += 1
            continue
        if llm_value not in (None, ""):
            continue  # 이미 판단된 행은 재호출하지 않음 (비용 절약)

        verdict = run_llm_judgement(claim_text)
        ws.cell(row=r, column=col["llm_verifiable"], value=verdict)
        missing_llm += 1

    wb.save(WORKBOOK_PATH)

    if unlabeled:
        print(f"경고: claim_class/source_scope가 비어있는 행 {unlabeled}건은 건너뛰었습니다. 라벨링을 마저 채워주세요.")
    print(f"LLM 판단 {missing_llm}건 신규 실행, 결과를 {WORKBOOK_PATH}에 저장했습니다.")

    df = pd.read_excel(WORKBOOK_PATH, sheet_name="라벨링")
    df = df.dropna(subset=["claim_class", "source_scope", "llm_verifiable"])
    if df.empty:
        print("비교할 라벨이 아직 없습니다. 라벨링 후 다시 실행하세요.")
        return

    df["human_verifiable"] = df.apply(
        lambda r: human_verifiable(r["claim_class"], r["source_scope"]), axis=1
    )
    agree = (df["human_verifiable"] == df["llm_verifiable"]).mean()

    tp = ((df["human_verifiable"]) & (df["llm_verifiable"])).sum()
    fp = ((~df["human_verifiable"]) & (df["llm_verifiable"])).sum()
    fn = ((df["human_verifiable"]) & (~df["llm_verifiable"])).sum()
    tn = ((~df["human_verifiable"]) & (~df["llm_verifiable"])).sum()

    report = f"""# claim_class 라벨링 파일럿 결과

> `claim_pilot_sample.py`로 뽑은 {len(df)}건 표본에 대해, 사람이 매긴
> claim_class/source_scope(→ 검증시도 여부)와 LLM({NCP_MODEL})의
> 판단(문장을 단독 기사로 넣었을 때 claim을 추출하는지)을 비교한 결과.

## 요약

- 표본 수: {len(df)}
- 사람-LLM 일치율: {agree:.1%}
- 혼동행렬 (사람 기준 '검증시도' = Positive)

| | LLM: 검증가능 | LLM: 검증불가 |
|---|---|---|
| 사람: 검증시도 | TP={tp} | FN={fn} |
| 사람: 판단불가 | FP={fp} | TN={tn} |

## 다음 단계

- 일치율이 충분히 높으면(팀 기준 합의 필요) 200~300건으로 확대
- FN(사람은 검증 가능하다고 봤는데 LLM이 놓친 경우), FP(반대)를 각각 몇 건 열어보고
  원인이 프롬프트 문제인지 스키마 정의 자체의 모호함인지 구분할 것
"""
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"보고서 작성: {REPORT_PATH}")
    print(f"일치율: {agree:.1%} (TP={tp} FP={fp} FN={fn} TN={tn})")


if __name__ == "__main__":
    main()
