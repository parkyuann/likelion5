"""Export the L2 contract v3 review workbook for human adjudication.

The reviewer-facing columns hold judgement only.  Character offsets are not
authored here: ``l2_span_resolver`` derives them from ``source_span_text`` at
ingest time and rejects text that is absent or ambiguous.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HUMAN_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
REFERENCE_FILL = PatternFill("solid", fgColor="EDEDED")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

HUMAN_INPUT_COLUMNS = (
    "indicator_scopes_json",
    "source_regions_json",
    "period_contexts_json",
    "clause_value_boundaries_json",
    "dominant_region_decision",
    "label_provenance",
    "review_status",
    "reviewer_note",
)
CONTEXT_INPUT_COLUMNS = ("disagree_flag", "reviewer_note")
WIDE_COLUMNS = {
    "text",
    "title",
    "value_candidates_text",
    "value_candidate_span_ids",
    "indicator_scopes_json",
    "source_regions_json",
    "period_contexts_json",
    "clause_value_boundaries_json",
    "reviewer_note",
    "indicator_label",
}


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def _cell_value(value: Any) -> Any:
    """Render nested structures as JSON text; Excel cells hold scalars only."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _write_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    input_columns: tuple[str, ...],
) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        return
    header = list(rows[0])
    sheet.append(header)
    for index, name in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = (
            48 if name in WIDE_COLUMNS else 18
        )
    for row in rows:
        sheet.append([_cell_value(row.get(name, "")) for name in header])
    for row_index in range(2, len(rows) + 2):
        for index, name in enumerate(header, start=1):
            cell = sheet.cell(row=row_index, column=index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if name in input_columns:
                cell.fill = HUMAN_INPUT_FILL
            elif title == "기사별 문맥 117문장" and name in {
                "scope_id",
                "region_id",
                "indicator_label",
                "source_subtype",
            }:
                cell.fill = REFERENCE_FILL
    sheet.freeze_panes = "A2"


def _add_enum_validation(
    workbook: Workbook,
    sheet_title: str,
    column_name: str,
    allowed: list[str],
    row_count: int,
) -> None:
    sheet = workbook[sheet_title]
    header = [cell.value for cell in sheet[1]]
    if column_name not in header:
        return
    letter = get_column_letter(header.index(column_name) + 1)
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(allowed) + '"',
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{row_count + 1}")


def _guide_rows(contract: dict[str, Any]) -> list[tuple[str, str]]:
    return [
        ("계약 버전", contract["contract_version"]),
        ("상태", contract["artifact_status"]),
        ("검토 행", str(contract["review_rows"])),
        (
            "사유 분포",
            ", ".join(
                f"{key} {value}"
                for key, value in contract["review_reason_counts"].items()
            ),
        ),
        ("", ""),
        (
            "v3에서 달라진 점",
            "문자 위치(char offset)를 사람이 입력하지 않습니다. "
            "source_span_text만 원문에서 그대로 복사해 넣으면 "
            "ingest 단계에서 위치를 결정론적으로 계산합니다.",
        ),
        (
            "span 규칙",
            contract["span_resolution_rule"],
        ),
        (
            "같은 표현이 두 번 나올 때",
            "occurrence_index(0부터)를 넣거나 더 긴 span을 쓰세요. "
            "모호한 span은 오류로 중단되며 임의로 앞선 것을 고르지 않습니다.",
        ),
        (
            "경계 규칙",
            contract["boundary_reference_rule"],
        ),
        (
            "상속 scope",
            "attribution_type이 '앞에서 상속'이면 그 문장에 span이 없어도 "
            "됩니다. 정의한 문장의 scope_id를 참조하세요.",
        ),
        (
            "ID 규칙",
            f"scope {contract['scope_id_format']}, "
            f"region {contract['region_id_format']}. "
            f"유일 범위: {contract['id_uniqueness_scope']} 단위.",
        ),
        (
            "지배 region",
            contract["reference_integrity_rule"],
        ),
        (
            "자동확정 이의",
            contract["disagreement_rule"],
        ),
        (
            "확신이 없을 때",
            "빈칸으로 두고 reviewer_note에 이유를 쓰세요. 추측해서 채우면 "
            "gold가 오염됩니다.",
        ),
        (
            "suggestion 사용 금지",
            "자동제안 workbook은 GOLD가 아닙니다. 판단값을 복사하지 마세요. "
            "특히 제안된 span은 부정확한 경우가 있습니다.",
        ),
        (
            "source subtype",
            ", ".join(contract["source_region_subtypes"]),
        ),
    ]


def export_workbook(
    human_rows: list[dict[str, Any]],
    context_rows: list[dict[str, Any]],
    contract: dict[str, Any],
    output_path: Path,
) -> dict[str, Any]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "검토입력", human_rows, HUMAN_INPUT_COLUMNS)
    _write_sheet(
        workbook,
        "기사별 문맥 117문장",
        context_rows,
        CONTEXT_INPUT_COLUMNS,
    )
    guide = workbook.create_sheet("계약·가이드")
    guide.append(["항목", "내용"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 110
    for key, value in _guide_rows(contract):
        guide.append([key, value])
    for row in guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _add_enum_validation(
        workbook,
        "검토입력",
        "label_provenance",
        list(contract["label_provenance_values"]),
        len(human_rows),
    )
    _add_enum_validation(
        workbook,
        "검토입력",
        "review_status",
        ["미검토", "검토완료", "보류"],
        len(human_rows),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "output": str(output_path),
        "sheets": workbook.sheetnames,
        "review_rows": len(human_rows),
        "context_rows": len(context_rows),
        "contract_version": contract["contract_version"],
        "human_input_columns": list(HUMAN_INPUT_COLUMNS),
        "offset_columns_present": False,
    }


def export_suggestion_workbook(
    suggestion_rows: list[dict[str, Any]],
    output_path: Path,
) -> dict[str, Any]:
    """Export suggestions to a physically separate, clearly labelled file."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    notice = workbook.create_sheet("경고")
    notice.column_dimensions["A"].width = 110
    for line in (
        "이 파일은 GOLD가 아닙니다.",
        "자동 파생 suggestion이며 사람 검토 결과로 사용할 수 없습니다.",
        "판단값을 검토 workbook에 복사하지 마세요.",
        "제안된 span은 부정확한 경우가 있습니다 — 한 문장의 여러 지표가 같은 "
        "부분 토큰을 가리키는 degenerate 결과가 포함돼 있습니다.",
        "원문에서 직접 span 텍스트를 고르세요.",
    ):
        notice.append([line])
    for row in notice.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _write_sheet(workbook, "자동제안", suggestion_rows, ())
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "output": str(output_path),
        "sheets": workbook.sheetnames,
        "suggestion_rows": len(suggestion_rows),
        "is_gold": False,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--human-jsonl", type=Path, required=True)
    parser.add_argument("--context-jsonl", type=Path, required=True)
    parser.add_argument("--contract", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--suggestions-jsonl", type=Path)
    parser.add_argument("--suggestions-output", type=Path)
    args = parser.parse_args()
    result = export_workbook(
        _read_jsonl(args.human_jsonl),
        _read_jsonl(args.context_jsonl),
        json.loads(args.contract.read_text(encoding="utf-8")),
        args.output,
    )
    if args.suggestions_jsonl and args.suggestions_output:
        result["suggestions"] = export_suggestion_workbook(
            _read_jsonl(args.suggestions_jsonl),
            args.suggestions_output,
        )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
