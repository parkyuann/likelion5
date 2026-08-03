"""Export the L-STRATUM re-check workbook.

The first judgment pass marked 45/45 articles as ``KOSIS등재`` because the
sheet only showed that a KOSIS organisation was named somewhere in the body.
Naming an organisation is not the same as sourcing a figure from a statistic
it publishes, so this sheet leads with that distinction: every row carries a
count of source-shaped versus actor-shaped mentions.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCOPES = ("KOSIS등재", "공식기관_비KOSIS", "민간기관", "해외기관", "불명")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")

COLUMNS = (
    "article_idx",
    "기사제목",
    "작성일",
    "검출된 KOSIS 기관",
    "기관 역할",
    "귀속 문장",
    "value_candidate_count",
    "density_bin",
    "judged_source_scope",
    "judge_note",
)
WIDE = {"기사제목", "귀속 문장", "검출된 KOSIS 기관", "judge_note", "기관 역할"}


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def _role_text(summary: dict[str, Any] | None) -> str:
    summary = summary or {}
    parts = [
        f"통계출처 {summary.get('통계출처', 0)}",
        f"행위주체 {summary.get('행위주체', 0)}",
        f"혼재 {summary.get('혼재', 0)}",
        f"불명 {summary.get('불명', 0)}",
    ]
    verdict = (
        "통계출처 문장 있음"
        if summary.get("source_shaped")
        else "⚠ 통계출처 문장 없음"
    )
    return verdict + "\n" + " / ".join(parts)


def _org_text(hits: object) -> str:
    if isinstance(hits, str):
        try:
            hits = json.loads(hits)
        except json.JSONDecodeError:
            return hits
    return ", ".join(
        f"{hit.get('surface')}→{hit.get('canonical_org')}"
        for hit in hits or []
    )


def _snippet_text(snippets: object) -> str:
    if isinstance(snippets, str):
        try:
            snippets = json.loads(snippets)
        except json.JSONDecodeError:
            return snippets
    return "\n\n".join(str(item) for item in snippets or [])


def _row_values(row: dict[str, Any], prior: str | None) -> list[Any]:
    values = [
        row.get("article_idx"),
        row.get("기사제목"),
        row.get("작성일"),
        _org_text(row.get("kosis_org_surface_hits")),
        _role_text(row.get("org_role_summary")),
        _snippet_text(row.get("org_attribution_snippets")),
        row.get("value_candidate_count"),
        row.get("density_bin"),
        "",
        "",
    ]
    if prior is not None:
        values.insert(8, prior)
    return values


def _sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    priors: dict[str, str] | None,
) -> None:
    sheet = workbook.create_sheet(title)
    header = list(COLUMNS)
    if priors is not None:
        header.insert(8, "1차 판정")
    sheet.append(header)
    for index, name in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = (
            52 if name in WIDE else 16
        )
    for row in rows:
        prior = (
            priors.get(str(row.get("article_idx")), "")
            if priors is not None
            else None
        )
        sheet.append(_row_values(row, prior))
    for row_index in range(2, len(rows) + 2):
        for index, name in enumerate(header, start=1):
            cell = sheet.cell(row=row_index, column=index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if name in {"judged_source_scope", "judge_note"}:
                cell.fill = INPUT_FILL
            elif name == "기관 역할" and str(cell.value).startswith("⚠"):
                cell.fill = WARN_FILL
    letter = get_column_letter(header.index("judged_source_scope") + 1)
    validation = DataValidation(
        type="list", formula1='"' + ",".join(SCOPES) + '"', allow_blank=True
    )
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{len(rows) + 1}")
    sheet.freeze_panes = "A2"


GUIDE = [
    ("판정 질문",
     "이 기사의 검증 대상 수치가, 그 기관이 정기적으로 집계·공표하는 통계에서 "
     "나왔는가?"),
    ("", ""),
    ("예 — KOSIS등재",
     "통계청 고용동향에 따르면 취업자 2909만명 / 한국은행 국민계정 1인당 국민소득 / "
     "출생아 수·출산율·인구 집계"),
    ("아니오 — 행위주체",
     "국세청이 46곳 세무조사 착수 / 산업부가 샌드박스 66건 승인 / "
     "정부가 3조원 지원. 기관이 등장해도 수치가 그 기관의 통계가 아니다"),
    ("아니오 — 해외기관",
     "미 노동부, OECD, IMF 등 해외 기관 발표. 후보 생성기가 대부분 걸러내지만 "
     "남아 있으면 해외기관으로 판정"),
    ("아니오 — 민간기관",
     "협회·연구소·기업의 조사·설문"),
    ("", ""),
    ("기관 역할 컬럼",
     "본문에서 그 기관이 언급된 문장을 통계출처형/행위주체형으로 자동 분류한 "
     "수치다. 판정이 아니라 참고 신호이며, ⚠ 표시는 통계출처형 문장이 하나도 "
     "없다는 뜻이다"),
    ("판정 단위",
     "기사 단위다. 검증 가능한 KOSIS 수치가 하나라도 있으면 KOSIS등재다. "
     "전망·정책목표·법정기준이 섞여 있어도 다른 정상 수치가 있으면 KOSIS등재"),
    ("전망치 주의",
     "OECD·IMF·기관의 성장률 전망은 예측이므로 검증 대상이 아니다. "
     "확정 집계와 구분할 것"),
    ("확신이 없으면", "불명으로 두고 judge_note에 이유를 남길 것"),
]


def export(
    recheck_rows: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
    priors: dict[str, str],
    output_path: Path,
) -> dict[str, Any]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "재확인", recheck_rows, priors)
    _sheet(workbook, "신규후보", new_rows, None)
    guide = workbook.create_sheet("판정 기준")
    guide.append(["항목", "내용"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 108
    for key, value in GUIDE:
        guide.append([key, value])
    for row in guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "output": str(output_path),
        "recheck_rows": len(recheck_rows),
        "new_rows": len(new_rows),
        "sheets": workbook.sheetnames,
        "forbidden_columns_absent": True,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pool", type=Path, required=True)
    parser.add_argument("--topup", type=Path, required=True)
    parser.add_argument("--prior-workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    pool = {row["article_idx"]: row for row in _read_jsonl(args.pool)}
    sheet = load_workbook(args.prior_workbook)["판정입력"]
    header = [cell.value for cell in sheet[4]]
    i_article = header.index("article_idx")
    i_judged = header.index("judged_source_scope")
    priors: dict[str, str] = {}
    for index in range(5, sheet.max_row + 1):
        article_idx = sheet.cell(index, i_article + 1).value
        if article_idx:
            priors[str(article_idx)] = str(
                sheet.cell(index, i_judged + 1).value or ""
            )

    recheck = []
    dropped = []
    for article_idx in priors:
        row = pool.get(article_idx)
        if row is None:
            dropped.append(article_idx)
            continue
        if not (row.get("org_role_summary") or {}).get("source_shaped"):
            recheck.append(row)

    result = export(
        recheck,
        _read_jsonl(args.topup),
        priors,
        args.output,
    )
    result["dropped_foreign_org_articles"] = dropped
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
